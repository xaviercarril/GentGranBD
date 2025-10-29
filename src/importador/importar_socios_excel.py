import os
from dataclasses import dataclass
from datetime import datetime
from numbers import Integral, Real
from typing import Any, Callable, Optional

import pandas as pd  # lazy import
from pandas.errors import EmptyDataError, ParserError
from sqlalchemy.exc import IntegrityError

from database import SessionLocal
from controladores.socios import construir_socio_modelo


ProgressCb = Optional[Callable[[int, int], None]]  # (procesadas, total)
WarnCb = Optional[Callable[[int, str], None]]      # (fila_index, mensaje)
ErrorCb = Optional[Callable[[int, str], None]]     # (fila_index, mensaje)


@dataclass
class RowImportError:
    row_index: int  # 1-based row number as seen en el fitxer original (inclou capçalera)
    message: str
    raw_data: dict[str, Any]
    error_columns: list[str]


class RowImportValidationError(Exception):
    """Excepción marcada con columnas problemáticas para facilitar el reporte."""

    def __init__(self, message: str, columns: Optional[list[str]] = None):
        super().__init__(message)
        self.columns = columns or []


def _normalize_header(name: str) -> str:
    """Normaliza cabeceras para poder mapear alias fácilmente."""
    if name is None:
        return ""
    return "".join(ch for ch in str(name).strip().lower() if ch.isalnum())


def _build_column_lookup(columns) -> dict[str, str]:
    """
    Crea un diccionario que mapea cabeceras normalizadas a su nombre original.
    Si hay colisiones conserva la primera aparición (se asume que las columnas son únicas).
    """
    lookup: dict[str, str] = {}
    for col in columns:
        norm = _normalize_header(col)
        if norm and norm not in lookup:
            lookup[norm] = col
    return lookup


def _clean_str(value) -> str | None:
    """Convierte a str, limpia espacios y normaliza nulos (NaN/None)."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    lowered = text.lower()
    if lowered in {"nan", "none", "null"}:
        return None
    return text


def _is_row_empty(row) -> bool:
    """
    Detecta files sense informació útil per evitar validar-les.
    Es considera buida si tots els valors són nuls/descriptors en blanc.
    """
    return all(_clean_str(value) is None for _, value in row.items())


def _resolve_column(lookup: dict[str, str], *aliases: str) -> str | None:
    """Devuelve el nombre original de columna para cualquiera de los alias dados."""
    for alias in aliases:
        norm = _normalize_header(alias)
        if norm in lookup:
            return lookup[norm]
    return None


def _get_cell(row, lookup: dict[str, str], *aliases: str):
    """Recupera el valor de la primera cabecera disponible entre los alias proporcionados."""
    column = _resolve_column(lookup, *aliases)
    if column:
        val = row.get(column, "")
        if val is not None:
            return val
    return ""


def _parse_optional_int(value, campo: Optional[str] = None) -> int | None:
    """Normaliza IDs opcionales aceptando enters en qualsevol format habitual."""
    if value is None:
        return None
    if isinstance(value, Integral):
        return int(value)

    if isinstance(value, Real) and not isinstance(value, Integral):
        value_float = float(value)
        if value_float.is_integer():
            return int(value_float)
        raise RowImportValidationError(
            "El Nº soci ha de ser un enter (sense decimals).",
            columns=[campo] if campo else [],
        )

    text = _clean_str(value)
    if text is None:
        return None

    text_normalized = text.replace(",", ".")
    try:
        if "." in text_normalized:
            num_float = float(text_normalized)
            if not num_float.is_integer():
                raise ValueError
            return int(num_float)
        return int(text_normalized)
    except ValueError:
        raise RowImportValidationError(
            "El Nº soci ha de ser un enter.",
            columns=[campo] if campo else [],
        )


def _normalize_dni_value(value, campo: Optional[str] = None) -> str | None:
    """Devuelve el DNI como texto, permetent valors numèrics sense decimals."""
    if value is None:
        return None
    if isinstance(value, str):
        return _clean_str(value)
    if isinstance(value, Integral):
        return str(int(value))
    if isinstance(value, Real):
        value_float = float(value)
        if not value_float.is_integer():
            raise RowImportValidationError(
                "El DNI/NIE no pot ser decimal.",
                columns=[campo] if campo else [],
            )
        return str(int(value_float))

    text = _clean_str(value)
    return text


def importar_socios_desde_excel(
    ruta_archivo: str,
    on_progress: ProgressCb = None,
    on_warning: WarnCb = None,
    on_error: ErrorCb = None,
) -> tuple[int, list[RowImportError]]:
    """
    Importa socios desde un archivo Excel o CSV gestionando errores fila a fila.
    - Muestra progreso por callback (procesadas/total).
    - Emite warnings por campos faltantes no críticos (p.ej. email vacío).
    - Si ocurre un error duro (validación/duplicado), registra la fila y continúa.

    Retorna un tuple (creados, filas_erroneas), donde filas_erroneas contiene
    los datos originales de las filas que no pudieron importarse.
    """

    # Lectura robusta del archivo con mensajes claros para el usuario
    try:
        if not ruta_archivo or not isinstance(ruta_archivo, str):
            raise ValueError("No se ha proporcionado una ruta de archivo válida.")

        if not os.path.exists(ruta_archivo):
            raise FileNotFoundError(f"El archivo no se encontró: {ruta_archivo}")

        ext = os.path.splitext(ruta_archivo)[1].lower()
        if ext in {".xls", ".xlsx"}:
            df = pd.read_excel(ruta_archivo)
        elif ext == ".csv":
            # Permite detectar automáticamente delimitadores como ';' muy comunes en exports de Excel
            read_opts = {"sep": None, "engine": "python"}
            try:
                df = pd.read_csv(ruta_archivo, **read_opts)
            except UnicodeDecodeError:
                # Intento con una codificación común alternativa
                read_opts["encoding"] = "latin-1"
                df = pd.read_csv(ruta_archivo, **read_opts)
        else:
            raise ValueError("Formato de archivo no soportado. Usa .xlsx, .xls o .csv")

        if df.empty:
            raise EmptyDataError("El archivo no contiene datos.")

    except EmptyDataError:
        raise ValueError("El archivo está vacío o no contiene filas válidas.")
    except ParserError:
        raise ValueError("El archivo tiene un formato inválido. Revisa separadores y cabeceras.")
    except Exception as e:
        # Mensaje genérico para cualquier otro error de lectura, incluyendo detalles originales
        raise ValueError(f"No se pudo leer el archivo. Verifica que el formato sea válido y que el archivo no esté dañado. Detalles: {str(e)}")

    df = df.fillna("")  # Reemplaza NaN por cadenas vacías
    column_lookup = _build_column_lookup(df.columns)

    total = len(df)
    procesadas = 0
    creados = 0

    failed_rows: list[RowImportError] = []

    with SessionLocal() as db:
        db.autoflush = False
        for position, (_, row) in enumerate(df.iterrows(), start=0):
            row_number_display = position + 2  # Primera fila de dades és la 2 (la 1 és la capçalera)
            callback_index = row_number_display - 1  # UI suma +1, així veu la fila real
            fila_original = {col: row.get(col, "") for col in df.columns}
            datos: dict[str, Any] = {}

            dni_column = _resolve_column(
                column_lookup,
                "D.N.I.",
                "D.N.I",
                "DNI",
                "DNI/NIE",
                "NIF",
                "DOCUMENT",
                "DOCUMENTO",
                "NUM. DOCUMENT",
                "NUM DOCUMENT",
                "NUMERO DOCUMENT",
                "NUM DOC",
            )
            id_column = _resolve_column(column_lookup, "Nº SOCI", "NUM SOCI")
            nombre_column = _resolve_column(column_lookup, "NOM", "NOMBRE", "NOMBRE SOCI")
            fecha_alta_column = _resolve_column(column_lookup, "DATA D'ALTA", "DATA ALTA", "FECHA ALTA", "F. ALTA")
            fecha_baja_column = _resolve_column(column_lookup, "BAIXAS", "DATA BAIXA", "FECHA BAJA", "F. BAJA")

            try:
                if _is_row_empty(row):
                    continue

                dni_raw = row.get(dni_column, "") if dni_column else _get_cell(
                    row,
                    column_lookup,
                    "D.N.I.",
                    "D.N.I",
                    "DNI",
                    "DNI/NIE",
                    "NIF",
                    "DOCUMENT",
                    "DOCUMENTO",
                    "NUM. DOCUMENT",
                    "NUM DOCUMENT",
                    "NUMERO DOCUMENT",
                    "NUM DOC",
                )
                fecha_alta_raw = row.get(fecha_alta_column, "") if fecha_alta_column else _get_cell(
                    row, column_lookup, "DATA D'ALTA", "DATA ALTA", "FECHA ALTA", "F. ALTA"
                )
                fecha_baja_raw = row.get(fecha_baja_column, "") if fecha_baja_column else _get_cell(
                    row, column_lookup, "BAIXAS", "DATA BAIXA", "FECHA BAJA", "F. BAJA"
                )
                socio_id_raw = row.get(id_column, "") if id_column else _get_cell(
                    row, column_lookup, "Nº SOCI", "NUM SOCI"
                )
                socio_id_normalized = _parse_optional_int(socio_id_raw, campo=id_column or "Nº SOCI")
                if socio_id_normalized is None:
                    id_error_columns = []
                    if id_column:
                        id_error_columns = [id_column]
                    else:
                        for candidate in ("Nº SOCI", "NUM SOCI"):
                            if candidate in df.columns:
                                id_error_columns = [candidate]
                                break
                    raise RowImportValidationError(
                        "El Nº soci és obligatori.",
                        columns=id_error_columns,
                    )

                dni_normalized = _normalize_dni_value(dni_raw, campo=dni_column)

                datos = {
                    "id": socio_id_normalized,
                    "nombre": _clean_str(row.get(nombre_column, "")) or _clean_str(
                        _get_cell(row, column_lookup, "NOM", "NOMBRE", "NOMBRE SOCI")
                    ) or "",
                    "apellido1": _clean_str(_get_cell(row, column_lookup, "1r COGNOM", "COGNOM1", "PRIMER COGNOM", "APELLIDO1", "APELLIDO 1")) or "",
                    "apellido2": _clean_str(_get_cell(row, column_lookup, "2n COGNOM", "COGNOM2", "SEGON COGNOM", "APELLIDO2", "APELLIDO 2")),
                    "dniNie": _clean_str(dni_normalized),
                    "direccion": _clean_str(_get_cell(row, column_lookup, "ADREÇA", "DIRECCION", "ADREÇA 1", "DIRECCIÓ", "ADRECA")),
                    "telefonoFijo": _clean_str(_get_cell(row, column_lookup, "TELÈFON", "TELEFON", "TELEFONO")),
                    "telefonoMovil": _clean_str(_get_cell(row, column_lookup, "MÒBIL", "MOBIL", "TELÈFON MÒBIL", "MOVIL", "TELEFONO MOVIL")),
                    "email": _clean_str(_get_cell(row, column_lookup, "E-MAIL", "EMAIL", "CORREU", "CORREO")),
                    "grupoDifusion": _clean_str(_get_cell(row, column_lookup, "E", "GRUP DIFUSIO", "GRUPO DIFUSION")),
                    "fechaAlta": _parse_fecha(fecha_alta_raw, campo=fecha_alta_column),
                    "fechaBaja": _parse_fecha(fecha_baja_raw, optional=True, campo=fecha_baja_column),
                    "observaciones": _clean_str(_get_cell(row, column_lookup, "OBSERVACIONES", "OBSERVACIONS", "OBSERVACIONS 1")),
                }

                if not datos["dniNie"]:
                    raise RowImportValidationError(
                        "El DNI/NIE és obligatori.",
                        columns=[dni_column] if dni_column else [],
                    )

                # Warnings no críticos
                # if not datos["email"] and on_warning:
                #     on_warning(callback_index, "Email vacío")

                socio = construir_socio_modelo(datos)
                db.add(socio)
                db.commit()
                creados += 1
            except RowImportValidationError as e:
                db.rollback()
                fila_error_columnas = e.columns
                mensaje = str(e)
                if on_error:
                    on_error(callback_index, mensaje)
                failed_rows.append(
                    RowImportError(
                        row_index=row_number_display,
                        message=mensaje,
                        raw_data=fila_original,
                        error_columns=fila_error_columnas,
                    )
                )
            except IntegrityError as e:
                db.rollback()
                dni = datos.get("dniNie")
                num = datos.get("id")
                if num:
                    mensaje = f"Soci duplicat (Nº soci {num}). Ja existeix. "
                    fila_error_columnas = [id_column] if id_column else []
                elif dni:
                    mensaje = f"Soci duplicat (DNI/NIE {dni}). Ja existeix. "
                    fila_error_columnas = [dni_column] if dni_column else []
                else:
                    mensaje = "Registre duplicat. Ja existeix. "
                    fila_error_columnas = []
                if on_error:
                    on_error(callback_index, mensaje)
                failed_rows.append(
                    RowImportError(
                        row_index=row_number_display,
                        message=mensaje,
                        raw_data=fila_original,
                        error_columns=fila_error_columnas,
                    )
                )
            except Exception as e:
                db.rollback()
                raw_message = str(e)
                lowered = raw_message.lower()
                mensajes: list[str] = []
                columnas_detectadas: list[str] = []

                if any(token in lowered for token in ("dni", "nie")):
                    mensajes.append("El DNI/NIE falta o no té un format vàlid.")
                    if dni_column:
                        columnas_detectadas.append(dni_column)
                if any(token in lowered for token in ("nombre", "nom")):
                    mensajes.append("El nom és obligatori.")
                    if nombre_column:
                        columnas_detectadas.append(nombre_column)

                # Si no se ha identificado ningún mensaje concreto, usar el original.
                if not mensajes:
                    mensajes = [raw_message]

                # Elimina duplicados conservando orden.
                unique_msgs = list(dict.fromkeys(mensajes))
                msg = " ".join(unique_msgs)
                fila_error_columnas = [col for col in columnas_detectadas if col]

                if on_error:
                    on_error(callback_index, msg)
                failed_rows.append(
                    RowImportError(
                        row_index=row_number_display,
                        message=msg,
                        raw_data=fila_original,
                        error_columns=fila_error_columnas,
                    )
                )
            finally:
                procesadas += 1
                if on_progress:
                    on_progress(procesadas, total)

    return creados, failed_rows

def _parse_fecha(fecha_raw, optional: bool = False, campo: Optional[str] = None):
    if not fecha_raw:
        return None if optional else datetime.today().date()

    if isinstance(fecha_raw, datetime):
        return fecha_raw.date()

    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(fecha_raw).strip(), fmt).date()
        except ValueError:
            continue

    if optional:
        return None
    raise RowImportValidationError(
        f"Format de data no reconegut: {fecha_raw}",
        columns=[campo] if campo else [],
    )


def exportar_filas_erroneas(
    filas_erroneas: list[RowImportError],
    ruta_salida: str,
) -> None:
    """
    Exporta las filas que no se pudieron importar a un Excel o CSV.
    - Añade columnas auxiliares con el número de fila original (1-based) y el mensaje de error.
    - Si se genera Excel, aplica formato de texto rojo en las celdas problemáticas.
    """
    if not filas_erroneas:
        return

    if not ruta_salida:
        raise ValueError("No s'ha proporcionat una ruta de destí per guardar els errors.")

    base_columns: list[str] = list(filas_erroneas[0].raw_data.keys())
    registros: list[dict[str, Any]] = []
    for fallo in filas_erroneas:
        registro = {col: "" for col in base_columns}
        for col in base_columns:
            registro[col] = _prepare_export_value(fallo.raw_data.get(col, ""))
        registro["Fila origen"] = fallo.row_index
        registro["Error"] = fallo.message
        registros.append(registro)

    # Garantiza orden de columnas: primero metainformación, luego columnas originales.
    columnas = ["Fila origen", "Error"] + base_columns
    df = pd.DataFrame(registros)
    df = df[columnas]

    ext = os.path.splitext(ruta_salida)[1].lower()
    if ext == ".csv":
        df.to_csv(ruta_salida, index=False)
        return

    # Excel (xlsx, xlsm…)
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Errors")

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
    except ImportError:
        return

    try:
        wb = load_workbook(ruta_salida)
    except Exception:
        return

    ws = wb.active
    red_font = Font(color="FF0000")
    for row_offset, fallo in enumerate(filas_erroneas, start=2):  # 1-based header
        for col_name in fallo.error_columns:
            if not col_name or col_name not in df.columns:
                continue
            col_idx = df.columns.get_loc(col_name) + 1  # openpyxl is 1-based
            cell = ws.cell(row=row_offset, column=col_idx)
            cell.font = red_font
    wb.save(ruta_salida)


def _prepare_export_value(value: Any) -> Any:
    """Normaliza los valores para exportarlos a CSV/Excel sin objetos pandas específicos."""
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    if isinstance(value, float) and (value != value):  # NaN check
        return ""
    return value
