from datetime import date

from openpyxl import load_workbook

from exportador.excel_participantes import generar_excel_participantes_viaje_session
from models import (
    Actividad,
    EstadoInscripcion,
    EstadoPago,
    InscripcionSocio,
    Pago,
    Socio,
    TipoActividadEnum,
)


def test_generar_excel_participantes_viaje_incluye_no_socios_y_pagat(session, tmp_path):
    actividad = Actividad(
        nombre="Viatge Tarragona",
        tipo=TipoActividadEnum.VIATGE,
        numMaxAlumnos=30,
        precio_matricula=25,
    )
    session.add(actividad)
    session.commit()

    socio = Socio(
        dniNie="X123",
        nombre="Anna",
        apellido1="Garcia",
        telefonoMovil="600111222",
        fechaAlta=date.today(),
    )
    session.add(socio)
    session.commit()

    session.add(
        InscripcionSocio(
            socioID=socio.id,
            actividadID=actividad.id,
            fechaInscripcion=date(2026, 2, 1),
            estado=EstadoInscripcion.INSCRIT,
        )
    )
    no_socio = InscripcionSocio(
        socioID=None,
        actividadID=actividad.id,
        noSocioNombre="Persona",
        noSocioApellido1="Externa",
        noSocioDni="Y999",
        noSocioTelefono="699111222",
        fechaInscripcion=date(2026, 2, 2),
        estado=EstadoInscripcion.INSCRIT,
    )
    session.add(no_socio)
    session.commit()

    session.add(
        Pago(
            socioID=None,
            actividadID=actividad.id,
            inscripcionID=no_socio.id,
            fecha=date(2026, 2, 3),
            importe=25,
            estado=EstadoPago.PAGAT,
        )
    )
    session.commit()

    output = tmp_path / "participants.xlsx"
    generar_excel_participantes_viaje_session(session, actividad.id, str(output))

    wb = load_workbook(output)
    ws = wb["Participants"]
    assert ws["A1"].value == "Participants - Viatge Tarragona"
    assert ws["D3"].value == "DNI"
    assert ws["I5"].value == "Sí"
    assert ws["D5"].value == "Y999"
